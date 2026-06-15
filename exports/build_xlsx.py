#!/usr/bin/env python3
"""Bundle the two CSV exports into a single XLSX workbook with two tabs."""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_DIR = Path("/home/user/rr_landing_page1/exports")

SHEETS = [
    ("Unique Players (Dedup)", "unique_players.csv"),
    ("All Submissions", "all_submissions.csv"),
]

HEADER_FILL = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)


def add_sheet(wb: Workbook, title: str, csv_path: Path) -> None:
    ws = wb.create_sheet(title=title)
    with csv_path.open(encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    # Style header
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Auto-width — cap to 40 to keep workbook usable
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 8
        for cell in column_cells:
            if cell.value is not None:
                v = str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)  # delete the default empty sheet

    for title, csv_name in SHEETS:
        add_sheet(wb, title, OUT_DIR / csv_name)

    out_path = OUT_DIR / "players_consolidated.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
